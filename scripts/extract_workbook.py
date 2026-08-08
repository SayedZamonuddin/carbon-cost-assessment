#!/usr/bin/env python3
"""Extract the static datasets, default parameters and golden validation values from
Oxford_Simple_IAM_3_two-actor_v1.8.xlsx into JSON consumed by the React dashboard.

Run once (and again whenever the workbook is updated):

    python3 -m venv .venv && .venv/bin/pip install openpyxl
    .venv/bin/python scripts/extract_workbook.py

Outputs into shards-dashboard-react/src/data/model/:
    static.json    bundled input datasets + observation overlays
    defaults.json  the workbook's default policy / run parameters
    golden.json    cached model results, used by the engine's Jest tests
"""

import json
import os
import warnings

import openpyxl

warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, "Oxford_Simple_IAM_3_two-actor_v1.8.xlsx")
OUT = os.path.join(ROOT, "shards-dashboard-react", "src", "data", "model")

YEAR_START, YEAR_END = 1765, 2150
YEARS = list(range(YEAR_START, YEAR_END + 1))

wb = openpyxl.load_workbook(SRC, data_only=True)


# --------------------------------------------------------------------------- helpers
def table(sheet, name):
    """Return an Excel table as {header: [values]}."""
    ws = wb[sheet]
    rows = list(ws[ws.tables[name].ref])
    headers = [c.value for c in rows[0]]
    out = {h: [] for h in headers}
    for row in rows[1:]:
        for h, c in zip(headers, row):
            out[h].append(c.value)
    return out


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def align(tbl, year_col, value_col, years=YEARS, fill=0.0, hold_last=False):
    """Reindex a table column onto `years`. Missing years get `fill`, or the last
    known value when hold_last is set (matches the workbook's PRIMAP behaviour)."""
    lookup = {}
    for y, v in zip(tbl[year_col], tbl[value_col]):
        if isinstance(y, (int, float)):
            lookup[int(y)] = num(v)
    if not lookup:
        return [fill] * len(years)
    last_year = max(lookup)
    first_year = min(lookup)
    out, held = [], fill
    for y in years:
        key = min(max(y, first_year), last_year) if hold_last else y
        v = lookup.get(key)
        if v is None:
            v = held if hold_last else fill
        else:
            held = v
        out.append(v)
    return out


def align_optional(tbl, year_col, value_col, **kw):
    """align(), but tolerant of a column the workbook does not define."""
    if value_col not in tbl:
        return [0.0] * len(YEARS)
    return align(tbl, year_col, value_col, **kw)


def series(tbl, year_col, value_col, years=YEARS):
    """Reindex onto `years` keeping None for genuinely absent observations."""
    lookup = {
        int(y): num(v)
        for y, v in zip(tbl[year_col], tbl[value_col])
        if isinstance(y, (int, float))
    }
    return [lookup.get(y) for y in years]


# ------------------------------------------------------------------- emission scenarios
# Each scenario supplies world emissions: CO2 in GtC, CH4 in MtCH4, N2O in MtN2O-N2.
SCENARIO_TABLES = [
    ("NONE", "Example Data", "NONE_EM"),
    ("RCP3PD", "RCP3PD_EM", "RCP3PD_EM"),
    ("RCP45", "RCP45_EM", "RCP45_EM"),
    ("RCP6", "RCP6_EM", "RCP6_EM"),
    ("RCP85", "RCP85_EM", "RCP85_EM"),
    ("CUSTOM", "Example Data", "CUSTOM_EM"),
    ("PULSE_CO2", "Example Data", "PULSE_CO2_EM"),
    ("PULSE_CH4", "Example Data", "PULSE_CH4_EM"),
    ("PULSE_N2O", "Example Data", "PULSE_N2O_EM"),
    ("STEP_CO2", "Example Data", "STEP_CO2_EM"),
    ("STEP_CH4", "Example Data", "STEP_CH4_EM"),
    ("STEP_N2O", "Example Data", "STEP_N2O_EM"),
]

scenarios = {}
for name, sheet, tname in SCENARIO_TABLES:
    t = table(sheet, tname)
    year_col = "YEAR" if "YEAR" in t else "Year"
    if "CO2" in t:
        co2 = align(t, year_col, "CO2")
    else:  # RCP tables split fossil and land-use CO2
        fossil = align(t, year_col, "FossilCO2")
        other = align(t, year_col, "OtherCO2")
        co2 = [a + b for a, b in zip(fossil, other)]
    scenarios[name] = {
        "co2": co2,
        "ch4": align_optional(t, year_col, "CH4"),
        "n2o": align_optional(t, year_col, "N2O"),
    }

# The CUSTOM scenario as shipped in the workbook's INPUT DATA sheet. This is what the
# app loads by default and what an uploaded workbook/CSV replaces.
input_custom = table("INPUT DATA", "INPUT_CUSTOM")
custom_scenario = {
    "co2": align(input_custom, "YEAR", "CO2"),
    "ch4": align(input_custom, "YEAR", "CH4"),
    "n2o": align(input_custom, "YEAR", "N2O"),
    "otherRf": align(input_custom, "YEAR", "OTHER_RF"),
}
scenarios["CUSTOM"] = {k: custom_scenario[k] for k in ("co2", "ch4", "n2o")}

# ------------------------------------------------------------------ non-GHG forcing
# INPUT_OTHER_RF pulls <EMS_SCEN>_RF[<option>] - <EMS_SCEN>_RF[CO2CH4N2O_RF]: the model
# computes CO2/CH4/N2O forcing itself, so only the residual is supplied externally.
# Scenarios without an RF table (CUSTOM, NONE, pulses, steps) fall through to zero.
RF_OPTIONS = ["TOTAL_INCLVOLCANIC_RF", "TOTAL_ANTHRO_RF", "GHG_RF", "TOTAL_NAT_RF"]
other_rf = {}
for name in ("RCP3PD", "RCP45", "RCP6", "RCP85"):
    t = table(name + "_RF", name + "_RF")
    base = align_optional(t, "YEAR", "CO2CH4N2O_RF")
    other_rf[name] = {
        opt: [a - b for a, b in zip(align_optional(t, "YEAR", opt), base)]
        if opt in t
        else [0.0] * len(YEARS)
        for opt in RF_OPTIONS
    }

# TEMP_NAT (natural-only warming, used by the "2018 anthropogenic warming" figure) is
# always driven by RCP4.5's volcanic+solar residual, whatever scenario is selected.
rcp45_rf = table("RCP45_RF", "RCP45_RF")
natural_rf = [
    a - b
    for a, b in zip(
        align(rcp45_rf, "YEAR", "TOTAL_INCLVOLCANIC_RF"),
        align(rcp45_rf, "YEAR", "TOTAL_ANTHRO_RF"),
    )
]

# ------------------------------------------------------------------------ observations
primap = table("PRIMAP_EM", "PRIMAP_EM")
hadcrut = table("HadCRUT TEMPS", "HadCRUT")
cmip26 = table("CMIP5 TEMPS", "CMIP5_RCP3PD_TEMP")
cmip85 = table("CMIP5 TEMPS", "CMIP5_RCP85_TEMP")

static = {
    "years": YEARS,
    "scenarios": scenarios,
    "customScenario": custom_scenario,
    "otherRf": other_rf,
    "naturalRf": natural_rf,
    "primap": {
        "co2": align(primap, "Year", "CO2 %World", hold_last=True),
        "ch4": align(primap, "Year", "CH4 %World", hold_last=True),
        "n2o": align(primap, "Year", "N2O %World", hold_last=True),
    },
    "observations": {
        "hadcrut": series(hadcrut, "YEAR", "Rel_to_1850_1900"),
        "cmip5rcp26": series(cmip26, "Year", "MODEL_MEAN"),
        "cmip5rcp85": series(cmip85, "Year", "MODEL_MEAN"),
    },
}

# --------------------------------------------------------------------------- defaults
dash = wb["Clim Policy Dash"]
model_dash = wb["Clim Model Dash"]


def cells(ws, addrs):
    return [ws[a].value for a in addrs]


def actor(row):
    (start, p2035, p2060, p2100, cost50, cost100, maxA, participation, ch4Abate) = cells(
        ws=dash,
        addrs=["%s%d" % (c, row) for c in "BCDEFGHIJ"],
    )
    return {
        "startYear": start,
        "price2035": p2035,
        "price2060": p2060,
        "price2100": p2100,
        "cost50": cost50,
        "cost100": cost100,
        "maxAbatement": maxA,
        "participation": participation,
        "ch4Abatement": ch4Abate,
    }


defaults = {
    "policy": {"annexI": actor(6), "nonAnnexI": actor(11)},
    "damage": {"damageAt1C": dash["B15"].value, "damageAt6C": dash["C15"].value},
    "economy": {
        "gwpValueYear": dash["F15"].value,
        "gwp": dash["G15"].value,
        "initialGrowthRate": dash["H15"].value,
        "growthDecline": dash["I15"].value,
        "discountRate": dash["J15"].value,
        "discountYear": dash["K15"].value,
        "finalCostYear": dash["L15"].value,
    },
    "run": {
        "emissionsScenario": model_dash["A3"].value,
        "otherRfScenario": model_dash["B3"].value,
        "ecs": model_dash["C3"].value,
        "tcr": model_dash["D3"].value,
        "policyOn": model_dash["E3"].value == "ON",
    },
    "display": {
        "windowFrom": model_dash["B6"].value,
        "windowTo": model_dash["C6"].value,
        "refFrom": model_dash["B7"].value,
        "refTo": model_dash["C7"].value,
    },
}

# ----------------------------------------------------------------------------- golden
temp_calc = table("RF & Temp", "TEMP_CALC")
temp_split = table("RF & Temp", "TEMP_SPLIT")
co2_calc = table("CO2", "CO2_CALC")
ch4_calc = table("CH4", "CH4_CALC")
n2o_calc = table("N2O", "N2O_CALC")
works_anx = table("Data Annex I", "ClimWorksCalcAnxI")
works_non = table("Data Non-Annex I", "ClimWorksCalcNonAnxI")
works_world = table("Clim Policy Dash", "ClimWorksCalc")
temp_baseline = table("RF & Temp", "TEMP_BASELINE")
co2_baseline = table("CO2", "CO2_BASELINE")

golden = {
    "years": YEARS,
    "kpis": {
        "warming2100": dash["B21"].value,
        "damage2100": dash["C21"].value,
        "totalDamage": dash["F20"].value,
        "totalMitigation": dash["G20"].value,
        "totalCost": dash["H20"].value,
        "discountedDamage": dash["F21"].value,
        "discountedMitigation": dash["G21"].value,
        "discountedCost": dash["H21"].value,
        "tcre": model_dash["G3"].value,
        "warming2018": model_dash["G5"].value,
    },
    "series": {
        "temp": align(temp_calc, "YEAR", "TEMP"),
        "tempRel": align(temp_calc, "YEAR", "TEMP_rel"),
        "rfTotal": align(temp_calc, "YEAR", "RF_TOT"),
        "tempCo2": align(temp_split, "YEAR", "TEMP_CO2"),
        "tempCh4": align(temp_split, "YEAR", "TEMP_CH4"),
        "tempN2o": align(temp_split, "YEAR", "TEMP_N2O"),
        "tempNat": align(temp_split, "YEAR", "TEMP_NAT"),
        "co2Conc": align(co2_calc, "YEAR", "CONC"),
        "co2Rf": align(co2_calc, "YEAR", "RF"),
        "co2Ems": align(co2_calc, "YEAR", "EMS"),
        "co2EmsAnxI": align(co2_calc, "YEAR", "AnxI_EMS"),
        "co2EmsNonAnxI": align(co2_calc, "YEAR", "NonAnxI_EMS"),
        "co2CumEms": align(co2_calc, "YEAR", "CUM_EMS"),
        "co2Alpha": align(co2_calc, "YEAR", "alpha"),
        "ch4Conc": align(ch4_calc, "YEAR", "CONC"),
        "ch4Rf": align(ch4_calc, "YEAR", "RF"),
        "ch4Ems": align(ch4_calc, "YEAR", "EMS"),
        "ch4Lifetime": align(ch4_calc, "YEAR", "LIFETIME"),
        "n2oConc": align(n2o_calc, "YEAR", "CONC"),
        "n2oRf": align(n2o_calc, "YEAR", "RF"),
        "n2oEms": align(n2o_calc, "YEAR", "EMS"),
        "baselineTempRel": align(temp_baseline, "YEAR", "TEMP_rel")
        if "YEAR" in temp_baseline
        else align(temp_calc, "YEAR", "TEMP_rel"),
        "baselineCo2Conc": align(co2_baseline, "YEAR", "CONC")
        if "YEAR" in co2_baseline
        else None,
        "baselineCo2Ems": align(co2_baseline, "YEAR", "EMS_BASELINE")
        if "YEAR" in co2_baseline
        else None,
        "annexPrice": align(works_anx, "Year", "CO2_Price"),
        "annexAbatement": align(works_anx, "Year", "Abatement"),
        "annexMitCost": align(works_anx, "Year", "Mitigation_cost"),
        "annexGwp": align(works_anx, "Year", "GWP"),
        "nonAnnexPrice": align(works_non, "Year", "CO2_Price"),
        "nonAnnexAbatement": align(works_non, "Year", "Abatement"),
        "nonAnnexMitCost": align(works_non, "Year", "Mitigation_cost"),
        "worldAbatement": align(works_world, "Year", "Abatement"),
        "worldMitCost": align(works_world, "Year", "Mitigation_cost"),
        "worldDamageFrac": align(works_world, "Year", "Climate_damages"),
        "worldTotalDamage": align(works_world, "Year", "Total_damage"),
        "worldDiscDamage": align(works_world, "Year", "Discd_damage"),
        "worldDiscMitCost": align(works_world, "Year", "Discd_mit_cost"),
        "worldGwp": align(works_world, "Year", "GWP"),
        "baselineDamageFrac": align(works_world, "Year", "Baseline_damages"),
        "baselineTotalDamage": align(works_world, "Year", "Baseline_Total_damage"),
        "baselineDiscDamage": align(works_world, "Year", "Baseline_discd_damage"),
        "annexShare": align(works_world, "Year", "Annex I World CO2 fraction"),
    },
}
# TEMP_BASELINE / CO2_BASELINE tables have no YEAR column of their own; they are row
# aligned with TEMP_CALC, so index them positionally.
for key, tbl, col in [
    ("baselineTempRel", temp_baseline, "TEMP_rel"),
    ("baselineCo2Conc", co2_baseline, "CONC"),
    ("baselineCo2Ems", co2_baseline, "EMS_BASELINE"),
]:
    if "YEAR" not in tbl and col in tbl:
        golden["series"][key] = [num(v) for v in tbl[col]][: len(YEARS)]

os.makedirs(OUT, exist_ok=True)
for name, payload in [
    ("static.json", static),
    ("defaults.json", defaults),
    ("golden.json", golden),
]:
    path = os.path.join(OUT, name)
    with open(path, "w") as fh:
        json.dump(payload, fh)
    print("wrote %s (%.1f KB)" % (path, os.path.getsize(path) / 1024))
